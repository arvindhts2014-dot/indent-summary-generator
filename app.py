import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO

st.set_page_config(page_title="Indent Summary Generator", layout="centered")

st.title("⛽ Indent Summary Generator")

uploaded_file = st.file_uploader(
    "Upload Indent Status Excel File",
    type=["xlsx", "xlsm"]
)

if uploaded_file:

    try:

        wb = load_workbook(uploaded_file, data_only=True)
        ws = wb.active

        rows = list(ws.values)

        headers = rows[0]
        df = pd.DataFrame(rows[1:], columns=headers)

        # Required columns
        sales_col = "Sales Document"
        material_col = "Material"
        qty_col = "OrderQty"
        created_col = "Created On"
        time_col = "Time"

        for col in [sales_col, material_col, qty_col, created_col, time_col]:
            if col not in df.columns:
                st.error(f"Column not found: {col}")
                st.stop()

        # Latest date in file
        df[created_col] = pd.to_datetime(df[created_col], errors="coerce")
        latest_date = df[created_col].dt.date.max()

        tpt_docs = set()
        dot_docs = set()
        dealer_docs = set()

        tpt_after_1400 = set()
        dot_after_1400 = set()

        mg_qty = 0.0
        hsd_qty = 0.0

        col_index = {c: i + 1 for i, c in enumerate(headers)}

        for excel_row in range(2, ws.max_row + 1):

            sales_doc = ws.cell(
                excel_row,
                col_index[sales_col]
            ).value

            material = str(
                ws.cell(
                    excel_row,
                    col_index[material_col]
                ).value
            ).strip()

            qty = ws.cell(
                excel_row,
                col_index[qty_col]
            ).value

            try:
                qty = float(qty)
            except:
                qty = 0

            # MG Qty
            if material in ["16730", "17295"]:
                mg_qty += qty

            # HSD Qty
            if material in ["50700", "50800"]:
                hsd_qty += qty

            fill = ws.cell(excel_row, 1).fill

            fill_type = fill.fill_type
            color_index = None

            try:
                color_index = fill.fgColor.indexed
            except:
                pass

            # TPT
            if fill_type is None:
                tpt_docs.add(sales_doc)

            # DoT
            elif color_index == 47:
                dot_docs.add(sales_doc)

            # Dealer
            elif color_index == 50:
                dealer_docs.add(sales_doc)

            # After 14:00 calculation
            try:

                row_date = pd.to_datetime(
                    ws.cell(
                        excel_row,
                        col_index[created_col]
                    ).value
                ).date()

                raw_time = ws.cell(
                    excel_row,
                    col_index[time_col]
                ).value

                if hasattr(raw_time, "hour"):
                    hour = raw_time.hour
                else:
                    hour = pd.to_datetime(str(raw_time)).hour

                if row_date == latest_date and hour >= 14:

                    if fill_type is None:
                        tpt_after_1400.add(sales_doc)

                    elif color_index == 47:
                        dot_after_1400.add(sales_doc)

            except:
                pass

        total_indents = (
            len(tpt_docs)
            + len(dot_docs)
            + len(dealer_docs)
        )

        report = f"""
Indent Summary Report
---------------------

TPT Indents                : {len(tpt_docs)}
(TPT Indents after 14:00   : {len(tpt_after_1400)})

DoT Indents                : {len(dot_docs)}
(DoT Indents after 14:00   : {len(dot_after_1400)})

Dealer Indents             : {len(dealer_docs)}

Total Indents              : {total_indents}

MG Qty (16730+17295)       : {mg_qty:.2f} KL

HSD Qty (50700+50800)      : {hsd_qty:.2f} KL
"""

        st.subheader("Indent Summary Report")

        st.code(report)

        st.download_button(
            "Download Report",
            report,
            file_name="Indent_Summary_Report.txt",
            mime="text/plain"
        )

    except Exception as e:
        st.error(f"Error: {str(e)}")
